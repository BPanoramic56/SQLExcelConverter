from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
import SQLHandler
import openpyxl
import configparser


config = configparser.ConfigParser()
config.read("configurations.ini")

TableName = config["Database Information"]["Table"]

Mode = config["Type Information"]["Mode"]

def setup():
    workbook = Workbook()
    sheet = workbook.active
    server_connection = SQLHandler.create_server_connection()

    return workbook, sheet, server_connection

def add_headers(server_connection, sheet):
    headers = list()

    for header in SQLHandler.get_headers(server_connection):
        headers.append(header['Field'])

    sheet.append(headers)
    return len(headers)

def add_rows(server_connection, sheet):

    for row in SQLHandler.get_all(server_connection):
        info = list()
        for header in SQLHandler.get_headers(server_connection):
            info.append(row[header["Field"]])
        sheet.append(info)

def create_table(server_connection, header_count):
    row_count = SQLHandler.get_row_count(server_connection)[0]['COUNT(*)']

    table = Table(displayName=f"{TableName}", ref=f"A1:{openpyxl.utils.get_column_letter(header_count)}{row_count}")

    style = TableStyleInfo(name=f"{TableName}", showFirstColumn=True,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    return table

def save_file(sheet, table, workbook):
    sheet.add_table(table)
    workbook.save(f"/Users/brunogomespascotto/Downloads/{TableName}.xlsx")

if __name__ == "__main__":
    if Mode == "SQL-Excel":
        workbook, sheet, server_connection = setup()
        header_count = add_headers(server_connection, sheet)
        add_rows(server_connection, sheet)
        table = create_table(server_connection, header_count)
        save_file(sheet, table, workbook)

    elif Mode == "Excel-SQL":
        server_connection = SQLHandler.create_server_connection()
        index = 1
        while SQLHandler.ensure_table_existence(server_connection, TableName):
            TableName = f"{TableName}{index}"
            index += 1
        Workbook = load_workbook(filename = config["Excel Information"]["Workbook"])
        Sheet = Workbook.worksheets[0]

        # print(Sheet.tables.values[0].ref)

        SQLHandler.create_table(server_connection, TableName, list(next(Sheet.iter_rows(min_row=1, max_row=1, values_only=True))))

        for row in Sheet.iter_rows(values_only=True):
            # print(list(row))
            SQLHandler.add_row(server_connection, TableName, list(row))